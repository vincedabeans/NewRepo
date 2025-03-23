import tkinter as tk
from pygame import mixer
import openpyxl as xl


root = tk.Tk()
root.title("Auto Pricing 2")
root.geometry("560x450")
root.config(bg="gray")
mixer.init()


cost_dict = {
    "short": {"text":{"colored": 7,
                      "partial": 5,
                      "grayscale": 3},
              "image":{"colored": 10,
                      "partial": 7,
                      "grayscale": 5},
              "both":{"colored": 9,
                      "partial": 6,
                      "grayscale": 3}
              },

    "long": {"text":{"colored": 9,
                      "partial": 7,
                      "grayscale": 5},
              "image":{"colored": 15,
                      "partial": 9,
                      "grayscale": 7},
              "both":{"colored": 12,
                      "partial": 8,
                      "grayscale": 5}
              },

    "a4": {"text": {"colored": 8,
                      "partial": 6,
                      "grayscale": 4},
             "image": {"colored": 12,
                       "partial": 8,
                       "grayscale": 6},
             "both": {"colored": 10,
                      "partial": 7,
                      "grayscale": 4}
             }

}



font_list = [
    ["Helvetica", 12],
    ["Arial", 14],
    ["Times New Roman", 16],
    ["Courier", 18],
    ["Verdana", 17],
    ["Comic Sans MS", 15],
    ["Tahoma", 24],
    ["Georgia", 26],
    ["Impact", 28],
    ["Lucida Console", 15]
]




def calculation(paper_type, print_type, color_type, number_of_pages):
    cost = cost_dict.get(paper_type, {}).get(print_type, {}).get(color_type, {})
    return cost * number_of_pages

# paper_cd = 'short'
# print_cd = 'both'
# color_cd = 'colored'
# number_pages = 50
# result = calculation(paper_cd, print_cd, color_cd, number_pages)
# print(result)
# until here perfectly working


#here makikita yung total price ng pinaprint
label = tk.Label(root, bg="light gray", font=("Arial", 20 ), width=34, height=2)
label.place(x=5, y=5)


wb = xl.load_workbook("Book1.xlsx")
ws = wb['Sheet1']



#I need to create a text_entry for the row so that we can specify the row in the GUI without having to edit thru IDE
next_row_entry = tk.Entry(root, width= 5, font=font_list[0], bg='light gray')
next_row_entry.place(x=300, y=280)


# from the text_entry, I need to get the int value from there and then use a function for it to work and specify the row
def add_to_excel():
    try:
        next_row = int(next_row_entry.get())
        pages = number_pages_entry.get(1.0, tk.END).strip()
        txt = text_result.get(1.0, tk.END).split()


        ws[f'A{next_row}'] = pages
        ws[f'B{next_row}'] = txt[0] if len(txt) > 0 else "" #if checks that len 0 has value if not then it will show nothing
        ws[f'C{next_row}'] = txt[1] if len(txt) > 1 else "" #checks if the len 1 has value if not then it will show nothing
        ws[f'D{next_row}'] = txt[2] if len(txt) > 2 else "" #basically to avoid error of not having enough value

        # warning for when the excel is open
        try:
            wb.save("book1.xlsx")
            label.config(text="Data saved successfully.")
        except PermissionError:
            label.config(text="Please close the Excel file before saving.")
    except(ValueError, KeyError):
        label.config(text="Invalid input. Please check your entries.")

# button for adding to excel
append_button = tk.Button(root, text="-->", bg="powder blue", width=10, height=2, command= lambda: add_to_excel())
append_button.place(x=360, y=280)



def add_to_calculation(btns):
    try:
        text_result.insert(2.0, btns)
        result = text_result.get(1.0, tk.END).strip()
        values = result.split()

        if len(values) >= 3:
            paper_type = values[0]
            print_type = values[1]
            color_type = values[2]

            number_of_pages = int(number_pages_entry.get(1.0, tk.END).strip())


            total_cost = calculation(paper_type, print_type, color_type, number_of_pages)


            label.config(text=f"Your total is:  P{total_cost} 🤞")


            #print(f"{paper_type} {print_type} {color_type} {number_of_pages}: {total_cost}")

    except(ValueError, KeyError):
        label.config(text="INVALID ENTRY😤😝")



def text_label_questions(container, side_pos, updown_pos, text):
    tk.Label(container, text=text, font=font_list[4], fg="Black", bg="Gray").place(x=side_pos, y=updown_pos)

text_label_questions(root, text="1. Paper Type: ", side_pos=30, updown_pos=140)
text_label_questions(root, text="2. Print Type: ", side_pos=30, updown_pos=185)
text_label_questions(root, text="3. Color Type: ", side_pos=30, updown_pos=230)

#/////////////////////////////////buttons for the calculations of papers ////////////////////////////////////////////

#button for paper type = short, long, a4
button1 = tk.Button(root, text="SHORT", fg="black", bg='pink', width=8, height=2, command= lambda: add_to_calculation("short "))
button1.place(x=300, y=140)

#y is vertical, the higher value it will go down

button11 = tk.Button(root, text="LONG", fg="black", bg='pink', width=8, height=2, command= lambda: add_to_calculation("long "))
button11.place(x=370, y=140)

button12 = tk.Button(root, text="A4", fg="black", bg='pink', width=8, height=2, command= lambda: add_to_calculation("a4 "))
button12.place(x=440, y=140)


#button for print type like = text, image, text ang image
button2 = tk.Button(root, text="TEXT", fg="black", bg='light blue', width=8, height=2, command= lambda: add_to_calculation("text "))
button2.place(x=300, y=185)

button21 = tk.Button(root, text="IMAGE", fg="black", bg='light blue', width=8, height=2,command= lambda: add_to_calculation("image "))
button21.place(x=370, y=185)

button22 = tk.Button(root, text="BOTH", fg="black", bg='light blue', width=8, height=2,command= lambda: add_to_calculation("both "))
button22.place(x=440, y=185)


#button for color type = none, partial, colored
button3 = tk.Button(root, text="GRAYSCALE", fg="black", bg='pink', width=8, height=2 ,command= lambda: add_to_calculation("grayscale "))
button3.place(x=300, y=230)

button31 = tk.Button(root, text="PARTIAL", fg="black", bg='pink', width=8, height=2,command= lambda: add_to_calculation("partial "))
button31.place(x=370, y=230)

button32 = tk.Button(root, text="COLORED", fg="black", bg='pink', width=8, height=2,command= lambda: add_to_calculation("colored "))
button32.place(x=440, y=230)


#kung ilang pages na ipiprint
number_pages_entry = tk.Text(root, bg="light blue", width=14, height=1 ,  font=font_list[5], fg="black")
number_pages_entry.place(x=300, y=100)


#dito makikita pinindot na buttons
text_result = tk.Text(root, font=[8], fg="blue", width=20, height=1, bg="white")
text_result.place(x=100, y=100)
#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////

def clear_textfield():
    try:
        text_result.delete(1.0, tk.END)
        number_pages_entry.delete(1.0, tk.END)
        label.config(text="")

    except(KeyError, ValueError):
        print("Engkkk")

clear_btn = tk.Button(root, text="C", font=font_list[9], width=5, bg="red", command=clear_textfield)
clear_btn.place(x=480, y=100)


def play_music():
    mixer.music.load('QuackSoundEffect.mp3')
    mixer.music.play()


def master_button():
    clear_textfield()
    play_music()


clear_btn.config(command=master_button)


root.mainloop()